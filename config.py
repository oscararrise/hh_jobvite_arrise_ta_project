import os
import requests
import webbrowser
from http.server import BaseHTTPRequestHandler, HTTPServer
from urllib.parse import urlparse, parse_qs
import traceback
import time
import hashlib
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime
import cyrtranslit
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import base64

from statics import (
    JOBVITE_HEADERS,
    HH_HEADERS,
    JOBVITE_JOB_URL,
    JOBVITE_CANDIDATE_URL,
    HH_EMPLOYER_ID,
    HH_CLIENT_ID,
    HH_CLIENT_SECRET,
    HH_REDIRECT_URI,
    HH_AUTH_URL,
    HH_TOKEN_URL,
    HH_API_URL,
    HH_ACCESS_TOKEN,
    HH_USER_AGENT,
    NAUKRI_SECRET_KEY,
    NAUKRI_ACCESS_KEY,
    POWER_AUTOMATE_REPORT_WEBHOOK_URL,
)


# =============================================================================
# BACKWARD COMPATIBILITY
# =============================================================================


headers = JOBVITE_HEADERS
headers_obtain_managers = HH_HEADERS

jobvite_url = JOBVITE_JOB_URL
url_insert_candidate = JOBVITE_CANDIDATE_URL

employer_id = HH_EMPLOYER_ID
client_id = HH_CLIENT_ID
client_secret = HH_CLIENT_SECRET
redirect_uri = HH_REDIRECT_URI

auth_url = HH_AUTH_URL
token_url = HH_TOKEN_URL
api_url = HH_API_URL

auth_code = None

secret_key = NAUKRI_SECRET_KEY
access_key = NAUKRI_ACCESS_KEY


# =============================================================================
# REPORT CONFIG
# =============================================================================

REPORTS_DIR = "reports"
REPORT_FILE_NAME = "jobvite_hhru_publication_report.xlsx"
REPORT_FILE_PATH = os.path.join(REPORTS_DIR, REPORT_FILE_NAME)

DETAIL_SHEET = "Posting Details"
SUMMARY_REQ_SHEET = "Summary Requisitions"
SUMMARY_EXEC_SHEET = "Summary Execution"


# =============================================================================
# GENERAL HELPERS
# =============================================================================

def now_str() -> str:
    try:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    except Exception as e:
        print(f"Error in now_str: {e}\n{traceback.format_exc()}")
        return ""


def ensure_reports_dir() -> None:
    try:
        os.makedirs(REPORTS_DIR, exist_ok=True)
    except Exception as e:
        print(f"Error in ensure_reports_dir: {e}\n{traceback.format_exc()}")


def to_latin_safe(value: Optional[str]) -> Optional[str]:
    try:
        if not value:
            return None
        return cyrtranslit.to_latin(str(value), "ru")
    except Exception:
        return value


def create_keys(job_title: str, job_id: str) -> Tuple[str, str]:
    try:
        signature = hashlib.md5(f"{secret_key}{access_key}".encode()).hexdigest()
        en_access_key = hashlib.md5(
            f"{secret_key}{job_title}{job_id}{access_key}".encode()
        ).hexdigest()

        return signature, en_access_key

    except Exception as e:
        print(f"Error in create_keys: {e}\n{traceback.format_exc()}")
        return "", ""


def clean_payload(data: Any) -> Any:
    try:
        if isinstance(data, dict):
            cleaned = {}
            for k, v in data.items():
                v = clean_payload(v)
                if v in ("", None, [], {}):
                    continue
                cleaned[k] = v
            return cleaned

        if isinstance(data, list):
            cleaned = [clean_payload(x) for x in data]
            return [x for x in cleaned if x not in ("", None, [], {})]

        return data

    except Exception as e:
        print(f"Error in clean_payload: {e}\n{traceback.format_exc()}")
        return data


def clean_none(data: Any) -> Any:
    try:
        if isinstance(data, dict):
            return {
                k: clean_none(v)
                for k, v in data.items()
                if v is not None and v != "" and v != [] and v != {}
            }

        if isinstance(data, list):
            return [
                clean_none(item)
                for item in data
                if item not in [None, "", [], {}]
            ]

        return data

    except Exception as e:
        print(f"Error in clean_none: {e}\n{traceback.format_exc()}")
        return data


# =============================================================================
# JOBVITE
# =============================================================================

def get_all_requisitions() -> List[Dict[str, Any]]:
    try:
        start = 1
        count = 500
        all_requisitions = []

        while True:
            params = {
                "jobStatus": "Open",
                "start": start,
                "count": count,
            }

            resp = requests.get(
                jobvite_url,
                headers=headers,
                params=params,
                timeout=30,
            )
            resp.raise_for_status()

            data = resp.json()
            requisitions = data.get("requisitions", [])

            if not requisitions:
                break

            all_requisitions.extend(requisitions)

            if len(requisitions) < count:
                break

            start += count

        return all_requisitions

    except Exception as e:
        print(f"Error in get_all_requisitions: {e}\n{traceback.format_exc()}")
        return []


def create_jobvite_candidate(
    candidate_data: Dict[str, Any],
    requisition_id: str,
) -> Dict[str, Any]:
    try:
        payload = build_jobvite_candidate_payload(candidate_data, requisition_id)

        response = requests.post(
            url_insert_candidate,
            headers=headers,
            json=payload,
            timeout=30,
        )

        try:
            response_json = response.json()
        except Exception:
            response_json = {"raw_response": response.text}

        if response.status_code not in [200, 201]:
            return {
                "success": False,
                "status_code": response.status_code,
                "message": str(response_json),
                "response_json": response_json,
                "response_id": None,
            }

        response_id = None

        if isinstance(response_json, dict):
            response_id = (
                response_json.get("id")
                or response_json.get("candidateId")
                or response_json.get("applicationId")
            )

        return {
            "success": True,
            "status_code": response.status_code,
            "message": "Candidate created successfully",
            "response_json": response_json,
            "response_id": response_id,
        }

    except Exception as e:
        print(f"Error in create_jobvite_candidate: {e}\n{traceback.format_exc()}")
        return {
            "success": False,
            "status_code": 0,
            "message": f"{e}",
            "response_json": {},
            "response_id": None,
        }


def map_gender(gender_value: Optional[str]) -> Optional[str]:
    try:
        if not gender_value:
            return None

        gender_value = gender_value.strip().lower()

        mapping = {
            "male": "Male",
            "female": "Female",
            "undefined": "Undefined",
        }

        return mapping.get(gender_value, "Undefined")

    except Exception as e:
        print(f"Error in map_gender: {e}\n{traceback.format_exc()}")
        return None


def map_work_history(
    experience_list: Optional[List[Dict[str, Any]]],
) -> List[Dict[str, Any]]:
    try:
        work_history = []

        for exp in experience_list or []:
            item = {}

            start = exp.get("start")
            end = exp.get("end")

            if exp.get("position"):
                item["title"] = to_latin_safe(exp.get("position"))

            if exp.get("company"):
                item["companyName"] = to_latin_safe(exp.get("company"))

            if start:
                parts = start.split("-")
                if len(parts) >= 2:
                    item["startYear"] = parts[0]
                    item["startMonth"] = parts[1]

            if end:
                parts = end.split("-")
                if len(parts) >= 2:
                    item["endYear"] = parts[0]
                    item["endMonth"] = parts[1]

            if item:
                work_history.append(item)

        return work_history

    except Exception as e:
        print(f"Error in map_work_history: {e}\n{traceback.format_exc()}")
        return []


def build_comments(candidate_data: Dict[str, Any]) -> str:
    try:
        parts = []

        field_order = [
            ("resume_id", candidate_data.get("resume_id")),
            ("vacancy_id", candidate_data.get("vacancy_id")),
            ("negotiation_id", candidate_data.get("negotiation_id")),
            ("title", candidate_data.get("title")),
            ("city", candidate_data.get("city")),
            ("experience_months", candidate_data.get("experience_months")),
            ("score", candidate_data.get("score")),
            ("applied_at", candidate_data.get("applied_at")),
            ("updated_at", candidate_data.get("updated_at")),
            ("resume_url", candidate_data.get("resume_url")),
            ("alternate_resume_url", candidate_data.get("alternate_resume_url")),
        ]

        for key, value in field_order:
            if value not in [None, ""]:
                parts.append(f"{key}={value}")

        return "HH " + " | ".join(parts) if parts else "HH candidate"

    except Exception as e:
        print(f"Error in build_comments: {e}\n{traceback.format_exc()}")
        return "HH candidate"


def build_jobvite_candidate_payload(
    candidate_data: Dict[str, Any],
    requisition_id: str,
) -> Dict[str, Any]:
    try:
        payload = {
            "sendEmail": "false",
            "sendOFCCPEmail": "false",
            "candidate": {
                "email": candidate_data.get("email"),
                "firstName": to_latin_safe(candidate_data.get("first_name")),
                "lastName": to_latin_safe(candidate_data.get("last_name")),
                "mobile": candidate_data.get("phone"),
                "city": candidate_data.get("city"),
                "workHistory": map_work_history(candidate_data.get("experience")),
                "application": {
                    "workflowState": "New",
                    "requisitionId": str(requisition_id),
                    "source": "HH",
                    "sourceType": "JobBoard",
                    "gender": map_gender(candidate_data.get("gender")),
                    "comments": "auto-posted from HH | " + build_comments(candidate_data),
                    "resume": {
                        "name": "candidate_resume.pdf",
                        "contentByteArray": candidate_data.get("pdf_base64"),
                        "format": "ByteArray",
                    },
                },
            },
        }

        return clean_none(payload)

    except Exception as e:
        print(
            f"Error in build_jobvite_candidate_payload: {e}\n"
            f"{traceback.format_exc()}"
        )
        return {}


# =============================================================================
# HH AUTH
# =============================================================================

class CallbackHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        try:
            global auth_code

            parsed = urlparse(self.path)
            params = parse_qs(parsed.query)

            if "code" in params:
                auth_code = params["code"][0]

            self.send_response(200)
            self.end_headers()
            self.wfile.write(b"Authorization successful. You can close this page.")

        except Exception as e:
            print(f"Error in CallbackHandler.do_GET: {e}\n{traceback.format_exc()}")
            try:
                self.send_response(500)
                self.end_headers()
                self.wfile.write(b"Internal server error.")
            except Exception:
                pass

    def log_message(self, format, *args):
        return


def get_code() -> Optional[str]:
    try:
        full_auth_url = (
            f"{auth_url}?response_type=code"
            f"&client_id={client_id}"
            f"&redirect_uri={redirect_uri}"
            f"&force_login=true"
        )

        webbrowser.open(full_auth_url)
        server = HTTPServer(("localhost", 8080), CallbackHandler)

        while auth_code is None:
            server.handle_request()

        server.server_close()
        return auth_code

    except Exception as e:
        print(f"Error in get_code: {e}\n{traceback.format_exc()}")
        return None


def get_token(code: str) -> Optional[str]:
    try:
        payload = {
            "grant_type": "authorization_code",
            "client_id": client_id,
            "client_secret": client_secret,
            "code": code,
            "redirect_uri": redirect_uri,
        }

        response = requests.post(
            token_url,
            data=payload,
            timeout=30,
        )
        response.raise_for_status()

        return response.json().get("access_token")

    except Exception as e:
        print(f"Error in get_token: {e}\n{traceback.format_exc()}")
        return None


def get_me(token: str) -> Dict[str, Any]:
    try:
        headers_local = {
            "Authorization": f"Bearer {token}",
            "HH-User-Agent": HH_USER_AGENT,
            "Accept": "application/json",
        }

        response = requests.get(
            f"{api_url}/me",
            headers=headers_local,
            timeout=30,
        )
        response.raise_for_status()

        return response.json()

    except Exception as e:
        print(f"Error in get_me: {e}\n{traceback.format_exc()}")
        return {}


# =============================================================================
# HH API
# =============================================================================

def get_all_managers(
    BASE_URL: str,
    HEADERS: Dict[str, str],
) -> List[str]:
    try:
        response = requests.get(
            f"{BASE_URL}/employers/{employer_id}/managers",
            headers=HEADERS,
            timeout=30,
        )

        if response.status_code != 200:
            print(f"Error {response.status_code}: {response.text}")
            return []

        managers = response.json().get("items", [])
        manager_ids = [m["id"] for m in managers if m.get("id")]

        return manager_ids

    except Exception as e:
        print(f"Error in get_all_managers: {e}\n{traceback.format_exc()}")
        return []


def get_all_vacancies_by_managers(
    token: str,
    manager_ids: List[str],
    base_url: str,
) -> List[Dict[str, Any]]:
    try:
        headers_local = {
            "Authorization": f"Bearer {token}",
            "HH-User-Agent": HH_USER_AGENT,
            "Accept": "application/json",
        }

        all_vacancies = []

        for manager_id in manager_ids:
            page = 0

            while True:
                params = {
                    "manager_id": manager_id,
                    "page": page,
                    "per_page": 50,
                }

                response = requests.get(
                    f"{base_url}/employers/{employer_id}/vacancies/active",
                    headers=headers_local,
                    params=params,
                    timeout=30,
                )

                if response.status_code != 200:
                    print(
                        f"Error manager {manager_id} | "
                        f"Status {response.status_code}: {response.text}"
                    )
                    break

                data = response.json()
                items = data.get("items", [])
                all_vacancies.extend(items)

                if page >= data.get("pages", 1) - 1:
                    break

                page += 1
                time.sleep(0.3)

        return all_vacancies

    except Exception as e:
        print(
            f"Error in get_all_vacancies_by_managers: {e}\n"
            f"{traceback.format_exc()}"
        )
        return []


def _build_headers(token: str) -> Dict[str, str]:
    try:
        return {
            "Authorization": f"Bearer {token}",
            "HH-User-Agent": HH_USER_AGENT,
            "Accept": "application/json",
        }
    except Exception as e:
        print(f"Error in _build_headers: {e}\n{traceback.format_exc()}")
        return {}


def _safe_json(response: requests.Response) -> Any:
    try:
        return response.json()
    except Exception:
        try:
            return response.text
        except Exception as e:
            print(f"Error in _safe_json: {e}\n{traceback.format_exc()}")
            return None


def _request_with_retry(
    session: requests.Session,
    method: str,
    url: str,
    headers: Dict[str, str],
    params: Optional[Dict[str, Any]] = None,
    timeout: int = 30,
    retries: int = 3,
    sleep_seconds: float = 0.5,
) -> Any:
    try:
        last_error = None

        for attempt in range(1, retries + 1):
            try:
                response = session.request(
                    method=method,
                    url=url,
                    headers=headers,
                    params=params,
                    timeout=timeout,
                )

                if response.status_code == 200:
                    return response.json()

                last_error = f"HTTP {response.status_code}: {_safe_json(response)}"

                if response.status_code in (429, 500, 502, 503, 504):
                    time.sleep(sleep_seconds * attempt)
                    continue

                raise Exception(last_error)

            except requests.RequestException as exc:
                last_error = str(exc)

                if attempt < retries:
                    time.sleep(sleep_seconds * attempt)
                    continue

                raise Exception(
                    f"Request failed after {retries} attempts: {last_error}"
                )

        raise Exception(last_error or "Unknown request error")

    except Exception as e:
        print(f"Error in _request_with_retry: {e}\n{traceback.format_exc()}")
        raise


def _normalize_contact_value(value: Any) -> Optional[str]:
    try:
        if value is None:
            return None

        if isinstance(value, str):
            return value.strip() or None

        if isinstance(value, dict):
            for key in ("value", "formatted", "number", "url"):
                if value.get(key):
                    return str(value[key]).strip()

        value_str = str(value).strip()
        return value_str if value_str else None

    except Exception as e:
        print(f"Error in _normalize_contact_value: {e}\n{traceback.format_exc()}")
        return None


def _extract_contacts(full_resume: Dict[str, Any]) -> Dict[str, Any]:
    try:
        contacts = full_resume.get("contact") or full_resume.get("contacts") or []

        email = None
        phone = None
        telegram = None
        skype = None
        linkedin = None
        other_contacts = []

        for contact in contacts:
            if not isinstance(contact, dict):
                continue

            contact_type = contact.get("type")

            if isinstance(contact_type, dict):
                contact_type = contact_type.get("id")

            contact_type = str(contact_type).lower() if contact_type else ""
            value = _normalize_contact_value(contact.get("value"))

            if not value:
                continue

            if contact_type == "email" and not email:
                email = value
            elif contact_type in ["cell", "phone"] and not phone:
                phone = value
            elif contact_type == "telegram" and not telegram:
                telegram = value
            elif contact_type == "skype" and not skype:
                skype = value
            elif contact_type == "linkedin" and not linkedin:
                linkedin = value
            else:
                other_contacts.append({
                    "type": contact_type,
                    "value": value,
                })

        return {
            "email": email,
            "phone": phone,
            "telegram": telegram,
            "skype": skype,
            "linkedin": linkedin,
            "other_contacts": other_contacts,
        }

    except Exception as e:
        print(f"Error in _extract_contacts: {e}\n{traceback.format_exc()}")
        return {
            "email": None,
            "phone": None,
            "telegram": None,
            "skype": None,
            "linkedin": None,
            "other_contacts": [],
        }


def _get_full_resume(
    session: requests.Session,
    token: str,
    resume_url: Optional[str],
) -> Dict[str, Any]:
    try:
        if not resume_url:
            return {}

        headers_local = _build_headers(token)

        data = _request_with_retry(
            session=session,
            method="GET",
            url=resume_url,
            headers=headers_local,
        )

        return data if isinstance(data, dict) else {}

    except Exception as e:
        print(f"Error in _get_full_resume: {e}\n{traceback.format_exc()}")
        return {}


def _simplify_experience(
    experience_list: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    try:
        return [
            {
                "start": exp.get("start"),
                "end": exp.get("end"),
                "company": exp.get("company"),
                "position": exp.get("position"),
            }
            for exp in experience_list or []
        ]
    except Exception as e:
        print(f"Error in _simplify_experience: {e}\n{traceback.format_exc()}")
        return []


def _download_resume_pdf_base64(
    token: str,
    full_resume: Dict[str, Any],
) -> Optional[str]:
    try:
        download_data = full_resume.get("download") or {}
        pdf_data = download_data.get("pdf") or {}
        pdf_url = pdf_data.get("url")

        if not pdf_url:
            return None

        headers_cv = {
            "Authorization": f"Bearer {token}",
            "HH-User-Agent": HH_USER_AGENT,
            "Accept": "application/pdf",
        }

        response_cv = requests.get(
            pdf_url,
            headers=headers_cv,
            timeout=30,
        )
        response_cv.raise_for_status()

        return base64.b64encode(response_cv.content).decode("utf-8")

    except Exception as e:
        print(f"Error in _download_resume_pdf_base64: {e}\n{traceback.format_exc()}")
        return None


def get_top_candidates(
    token: str,
    vacancy_id: str,
    min_score: int = 84,
    per_page: int = 20,
    include_full_resume: bool = True,
) -> List[Dict[str, Any]]:
    try:
        headers_local = _build_headers(token)
        candidates: List[Dict[str, Any]] = []
        page = 0

        with requests.Session() as session:
            while True:
                try:
                    data = _request_with_retry(
                        session=session,
                        method="GET",
                        url=f"{api_url}/negotiations/response",
                        headers=headers_local,
                        params={
                            "vacancy_id": vacancy_id,
                            "page": page,
                            "per_page": per_page,
                        },
                    )
                except Exception as exc:
                    print(f"Error in negotiations page={page}: {exc}")
                    break

                if isinstance(data, list):
                    items = data
                    total_pages = None
                elif isinstance(data, dict):
                    items = data.get("items", []) or []
                    total_pages = data.get("pages")
                else:
                    print(f"Unexpected format in page={page}: {type(data)}")
                    break

                if not items:
                    break

                for item in items:
                    test_result = item.get("test_result") or {}
                    score = test_result.get("score")

                    if score is None or score < min_score:
                        continue

                    resume = item.get("resume") or {}
                    full_resume = {}

                    if include_full_resume and resume.get("url"):
                        full_resume = _get_full_resume(
                            session=session,
                            token=token,
                            resume_url=resume.get("url"),
                        )

                    contacts = _extract_contacts(full_resume)

                    area = full_resume.get("area") or resume.get("area") or {}

                    total_experience = (
                        full_resume.get("total_experience")
                        or resume.get("total_experience")
                        or {}
                    )

                    experience = (
                        full_resume.get("experience")
                        or resume.get("experience")
                        or []
                    )

                    experience_simple = _simplify_experience(experience)
                    pdf_base64 = _download_resume_pdf_base64(token, full_resume)

                    gender_data = full_resume.get("gender") or resume.get("gender") or {}

                    candidate = {
                        "vacancy_id": vacancy_id,
                        "negotiation_id": item.get("id"),
                        "resume_id": resume.get("id"),
                        "resume_url": resume.get("url"),
                        "alternate_resume_url": resume.get("alternate_url"),
                        "first_name": full_resume.get("first_name") or resume.get("first_name"),
                        "last_name": full_resume.get("last_name") or resume.get("last_name"),
                        "middle_name": full_resume.get("middle_name") or resume.get("middle_name"),
                        "title": full_resume.get("title") or resume.get("title"),
                        "email": contacts["email"],
                        "phone": contacts["phone"],
                        "age": full_resume.get("age") or resume.get("age"),
                        "gender": gender_data.get("id"),
                        "gender_name": gender_data.get("name"),
                        "city": area.get("name"),
                        "city_id": area.get("id"),
                        "experience_months": total_experience.get("months"),
                        "score": score,
                        "applied_at": item.get("created_at"),
                        "updated_at": item.get("updated_at"),
                        "experience": experience_simple,
                        "pdf_base64": pdf_base64,
                    }

                    candidates.append(candidate)

                if total_pages is not None:
                    if page >= total_pages - 1:
                        break
                else:
                    if len(items) < per_page:
                        break

                page += 1
                time.sleep(0.3)

        return candidates

    except Exception as e:
        print(f"Error in get_top_candidates: {e}\n{traceback.format_exc()}")
        return []


# =============================================================================
# EXCEL REPORT
# =============================================================================

def _autosize_columns(ws) -> None:
    try:
        for column_cells in ws.columns:
            max_length = 0
            column_index = column_cells[0].column
            column_letter = get_column_letter(column_index)

            for cell in column_cells:
                try:
                    value = "" if cell.value is None else str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    except Exception as e:
        print(f"Error in _autosize_columns: {e}\n{traceback.format_exc()}")


def _apply_header_style(ws) -> None:
    try:
        fill = PatternFill(fill_type="solid", fgColor="1F4E78")
        font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = fill
            cell.font = font

    except Exception as e:
        print(f"Error in _apply_header_style: {e}\n{traceback.format_exc()}")


def _ensure_workbook() -> Tuple[Any, Any]:
    try:
        ensure_reports_dir()

        if os.path.exists(REPORT_FILE_PATH):
            wb = load_workbook(REPORT_FILE_PATH)
        else:
            wb = Workbook()

        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
            default_sheet = wb["Sheet"]
            wb.remove(default_sheet)

        for sheet_name in [DETAIL_SHEET, SUMMARY_REQ_SHEET, SUMMARY_EXEC_SHEET]:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)

        return wb, wb[DETAIL_SHEET]

    except Exception as e:
        print(f"Error in _ensure_workbook: {e}\n{traceback.format_exc()}")
        raise


def _write_detail_sheet(
    ws,
    report_rows: List[Dict[str, Any]],
) -> None:
    try:
        headers_detail = [
            "execution_started_at",
            "posted_at",
            "requisition_id",
            "hh_vacancy_id",
            "hh_vacancy_name",
            "hh_vacancy_url",
            "candidate_first_name",
            "candidate_last_name",
            "candidate_email",
            "candidate_phone",
            "candidate_score",
            "candidate_city",
            "candidate_age",
            "candidate_resume_id",
            "candidate_resume_url",
            "candidate_negotiation_id",
            "jobvite_status_code",
            "jobvite_success",
            "jobvite_message",
            "jobvite_response_id",
        ]

        if ws.max_row == 1 and ws["A1"].value is None:
            ws.append(headers_detail)

        existing_headers = [cell.value for cell in ws[1]]

        if existing_headers != headers_detail:
            ws.delete_rows(1, ws.max_row)
            ws.append(headers_detail)

        for row in report_rows:
            ws.append([row.get(col) for col in headers_detail])

        _apply_header_style(ws)
        _autosize_columns(ws)

    except Exception as e:
        print(f"Error in _write_detail_sheet: {e}\n{traceback.format_exc()}")
        raise


def _read_all_detail_rows(ws) -> List[Dict[str, Any]]:
    try:
        if ws.max_row < 2:
            return []

        headers_row = [cell.value for cell in ws[1]]
        data_rows = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            data_rows.append(dict(zip(headers_row, row)))

        return data_rows

    except Exception as e:
        print(f"Error in _read_all_detail_rows: {e}\n{traceback.format_exc()}")
        return []


def _write_summary_req_sheet(
    ws,
    all_detail_rows: List[Dict[str, Any]],
) -> None:
    try:
        ws.delete_rows(1, ws.max_row)

        headers_summary = [
            "requisition_id",
            "hh_vacancy_id",
            "hh_vacancy_name",
            "candidates",
            "first_posted_at",
            "last_posted_at",
        ]

        ws.append(headers_summary)

        grouped = {}

        for row in all_detail_rows:
            key = (
                str(row.get("requisition_id", "")),
                str(row.get("hh_vacancy_id", "")),
                str(row.get("hh_vacancy_name", "")),
            )

            if key not in grouped:
                grouped[key] = {
                    "count": 0,
                    "first_posted_at": row.get("posted_at"),
                    "last_posted_at": row.get("posted_at"),
                }

            grouped[key]["count"] += 1

            current_posted_at = row.get("posted_at")

            if current_posted_at:
                if (
                    not grouped[key]["first_posted_at"]
                    or current_posted_at < grouped[key]["first_posted_at"]
                ):
                    grouped[key]["first_posted_at"] = current_posted_at

                if (
                    not grouped[key]["last_posted_at"]
                    or current_posted_at > grouped[key]["last_posted_at"]
                ):
                    grouped[key]["last_posted_at"] = current_posted_at

        sorted_items = sorted(
            grouped.items(),
            key=lambda x: (x[0][0], x[0][1], x[0][2]),
        )

        for (req_id, vacancy_id, vacancy_name), data in sorted_items:
            ws.append([
                req_id,
                vacancy_id,
                vacancy_name,
                data["count"],
                data["first_posted_at"],
                data["last_posted_at"],
            ])

        _apply_header_style(ws)
        _autosize_columns(ws)

    except Exception as e:
        print(f"Error in _write_summary_req_sheet: {e}\n{traceback.format_exc()}")
        raise


def _write_summary_execution_sheet(
    ws,
    stats: Dict[str, Any],
    report_rows: List[Dict[str, Any]],
) -> None:
    try:
        headers_execution = [
            "execution_started_at",
            "report_generated_at",
            "vacancies_found",
            "vacancies_with_req",
            "vacancies_without_req",
            "candidates_evaluated",
            "candidates_posted_successfully",
            "candidates_failed",
            "rows_written_to_detail",
        ]

        if ws.max_row == 1 and ws["A1"].value is None:
            ws.append(headers_execution)

        existing_headers = [cell.value for cell in ws[1]]

        if existing_headers != headers_execution:
            ws.delete_rows(1, ws.max_row)
            ws.append(headers_execution)

        ws.append([
            stats.get("execution_started_at"),
            now_str(),
            stats.get("vacancies_found", 0),
            stats.get("vacancies_with_req", 0),
            stats.get("vacancies_without_req", 0),
            stats.get("candidates_evaluated", 0),
            stats.get("candidates_posted_successfully", 0),
            stats.get("candidates_failed", 0),
            len(report_rows),
        ])

        _apply_header_style(ws)
        _autosize_columns(ws)

    except Exception as e:
        print(f"Error in _write_summary_execution_sheet: {e}\n{traceback.format_exc()}")
        raise

def send_excel_report_to_power_automate(
    file_path: str,
    execution_stats: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Sends the generated Excel report to a Power Automate HTTP trigger.

    The Excel file is converted to base64 and sent inside a JSON payload.
    Power Automate can then attach it to an email, save it to SharePoint,
    or continue any reporting workflow.
    """
    try:
        if not POWER_AUTOMATE_REPORT_WEBHOOK_URL:
            return {
                "success": False,
                "message": "POWER_AUTOMATE_REPORT_WEBHOOK_URL is not configured.",
                "status_code": None,
                "response": {},
            }

        if not os.path.exists(file_path):
            return {
                "success": False,
                "message": f"Report file does not exist: {file_path}",
                "status_code": None,
                "response": {},
            }

        with open(file_path, "rb") as file:
            file_base64 = base64.b64encode(file.read()).decode("utf-8")

        payload = {
            "file_name": os.path.basename(file_path),
            "file_base64": file_base64,
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "generated_at": now_str(),
            "execution_stats": execution_stats,
        }

        response = requests.post(
            POWER_AUTOMATE_REPORT_WEBHOOK_URL,
            json=payload,
            timeout=60,
        )

        try:
            response_json = response.json()
        except Exception:
            response_json = {"raw_response": response.text}

        if response.status_code not in [200, 201, 202]:
            return {
                "success": False,
                "message": "Power Automate flow rejected the request.",
                "status_code": response.status_code,
                "response": response_json,
            }

        return {
            "success": True,
            "message": "Report sent successfully to Power Automate.",
            "status_code": response.status_code,
            "response": response_json,
        }

    except Exception as e:
        print(
            f"Error in send_excel_report_to_power_automate: {e}\n"
            f"{traceback.format_exc()}"
        )
        return {
            "success": False,
            "message": str(e),
            "status_code": None,
            "response": {},
        }

def save_publication_report(
    report_rows: List[Dict[str, Any]],
    stats: Dict[str, Any],
) -> Dict[str, Any]:
    try:
        wb, detail_ws = _ensure_workbook()

        if report_rows:
            _write_detail_sheet(detail_ws, report_rows)
        else:
            if detail_ws.max_row == 1 and detail_ws["A1"].value is None:
                _write_detail_sheet(detail_ws, [])

        all_detail_rows = _read_all_detail_rows(detail_ws)

        summary_req_ws = wb[SUMMARY_REQ_SHEET]
        summary_exec_ws = wb[SUMMARY_EXEC_SHEET]

        _write_summary_req_sheet(summary_req_ws, all_detail_rows)
        _write_summary_execution_sheet(summary_exec_ws, stats, report_rows)

        wb.save(REPORT_FILE_PATH)

        return {
            "success": True,
            "file_path": REPORT_FILE_PATH,
            "rows_written": len(report_rows),
            "total_historical_rows": len(all_detail_rows),
        }

    except Exception as e:
        print(f"Error in save_publication_report: {e}\n{traceback.format_exc()}")
        return {
            "success": False,
            "file_path": REPORT_FILE_PATH,
            "rows_written": 0,
            "total_historical_rows": 0,
        }